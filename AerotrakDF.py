from openpyxl import load_workbook
import matplotlib.pyplot as plt
from datetime import datetime


class AerotrakDF(BaseDF):

    def __init__(self, file_path, correct_time_add_minutes=0):
        sheet = load_workbook(file_path.replace("\\", "/")).active
        data = {
            "date_time": [],
            "0.3": [],
            "0.5": [],
            "5.0": []
        }
        for row in sheet.iter_rows(min_row=14, min_col=2):
            timestamp = datetime.strptime(row[3].value, "%d/%m/%Y %H:%M:%S").timestamp()
            if correct_time_add_minutes != 0:
                timestamp += (correct_time_add_minutes * 60)
            data["date_time"].append(datetime.fromtimestamp(timestamp))
            data["0.3"].append(row[13].value)
            data["0.5"].append(row[16].value)
            data["5.0"].append(row[20].value)
        super().__init__(data)
        
    def show_plot(self):
        for particle_size, color in zip(self.keys()[1:], ('red', 'orange', 'green')):
            self.plot(x='date_time',y=particle_size,color=color, grid=True)
    #     self.plot(grid=True)
        plt.show()
     
    @staticmethod
    def compare(dataframes):
        for particle_size in dataframes[0].keys()[1:]:
            for df in dataframes:
                plt.plot(df["date_time"], df[particle_size])
            plt.show()
